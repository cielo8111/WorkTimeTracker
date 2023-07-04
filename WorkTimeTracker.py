from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font  
from tkinter import *
import tkinter as tk
import requests
from tkinter import messagebox
import pandas as pd
import os
import numpy as np
from tkinter import ttk
import tkinter.font
# import logging
import platform


def set_font(os):
    global font, font_log, font_search
    if platform.system == "Windows":
        font = tkinter.font.Font(os, family="メイリオ", size=10)
        font_log = tkinter.font.Font(os, family="メイリオ", size=8)
        font_search = tkinter.font.Font(os, family="メイリオ", size=7)
        print("윈도우즈용 폰트 세팅 / Font setting for Windows")

    else:
        font = tkinter.font.Font(os, family="メイリオ", size=13)
        font_log = tkinter.font.Font(os, family="メイリオ", size=11)
        font_search = tkinter.font.Font(os, family="メイリオ", size=10)
        print("맥 및 기타 운영체제의 폰트 세팅 / Font setting for Mac or other operating systems")


# 현재 날짜와 시간을 업데이트 / Update Current Date and Time
def update_datetime():
    global date_now, time_now
    date_now = datetime.now().strftime("%Y/%m/%d")
    time_now = datetime.now().strftime("%H:%M")
    return date_now, time_now


#로그창 이벤트 제어 / Control Event for Log Window
def handle_key_event(event):
    if(12==event.state and event.keysym=='c' ):
        return
    else:
        return "break"
    

# 라인봇 관련 / Line Bot Related
# 토큰 값 읽어오기 / Read Token Value
token = os.environ.get("TOKEN")
# 토큰 값이 있는지 확인 / Check if Token Value Exists
if token:
    # 토큰이 존재하면 사용하여 작업 수행 / Perform Operations if Token Exists
    print("토큰 값 / Token value:", token)
else:
    # 토큰이 존재하지 않을 경우 테스트 값 사용 / If the token does not exist, use a test value
    token = "eoXvQopvNSGia4ez8TP6IBUhEwzIs1TPctxPIXtcgft" #테스트용 : shion에게만 송신
    url = "https://notify-api.line.me/api/notify"
    print("토큰이 설정되어 있지 않음. 기본 값을 사용 / Token value is not set. Using default value.")


#라인봇 메시지 / Line Bot Message
def send_line_message(add_line_msg, *args):
    auth = {"Authorization": "Bearer " + token}
    content = {"message": add_line_msg.format(*args[:add_line_msg.count('{}')])}
    requests.post(url, headers=auth, data=content)


#날짜 유효성 검사 / Date Validation
def validate_date(date_text):
    try:
        datetime.strptime(date_text,"%Y/%m/%d")
        return True
    except ValueError:
        print("Incorrect data format({0}), should be YYYY/MM/DD".format(date_text))
        return False


#시스템 로그 / System Log
def info_and_system_log(message=None):
    update_datetime()
    if message:
        log_text_file = open("{}{}.txt".format(logPath, excel_username), "a")
        log_text_file.write(message.format(date_now, time_now))
        log_text_file.close()

        info_log.insert(tkinter.CURRENT, message.format(date_now, time_now))

    else:
        log_text_file = open("{}{}.txt".format(logPath, excel_username), "a")
        log_text_file.write("\n・ネットワークの接続が不安定です\n   ローカルにのみ保存します\n   {} {}\n".format(date_now, time_now))
        log_text_file.close()

        info_log.insert(tkinter.CURRENT, "\n・ネットワークの接続が不安定です\n   ローカルにのみ保存します\n   {} {}\n".format(date_now, time_now))
        return


def create_custom_label(message, x=None, y=None):
    label = tkinter.Label(window, bg=bg_var.get(), fg=fg_var.get(), font=font, text=message)   
    if x is not None and y is not None:
        label.place(x=x, y=y)
    return label


def create_button(text, command, x=None, y=None, w=None, h=None):
    button = tkinter.Button(window, highlightbackground="#323332", font=font, text=text, command=command) 
    if x is not None and y is not None:
        button.place(x=x, y=y, width=w, height=h)
    return button


def create_custom_entry(font_type):
    entry = tkinter.Entry(window, bg="#1E1E1E", fg="#ffffff", highlightbackground="#323332", font=font_type, insertbackground='#ffffff')
    return entry


def create_admin_pw():
    global admin_pw
    admin_pw = tkinter.Entry(window, show="*", bg="#1E1E1E", fg="#ffffff", highlightbackground="#323332",insertbackground='#ffffff')
    admin_pw.place(x=50, y=70, width=200, height=50)
    admin_pw.focus()
    return admin_pw

def save_workbook_to_paths(value1=None, value2=None):
    if value1 and value2:
        wb.save("{}{}.xlsx".format(value1,value2)) 
        wb.save("{}{}.xlsx".format(value1,value2)) 
    else:
        wb.save("{}{}.xlsx".format(uPath,excel_username)) 
        wb.save("{}{}.xlsx".format(bPath,excel_username))
        wb.save("{}{}.xlsx".format(ePath,excel_username))


#메인 GUI / Main GUI
window = tkinter.Tk() 
window.title("勤怠管理ツール") 
window.geometry("300x470") #
window.resizable(0,0) 
# window.attributes("-topmost", True)
window.configure(bg='#323332')
set_font(window)

info_log = tkinter.Text(window, highlightbackground="#323332", bg="#1E1E1E", fg="#ffffff", font=font_log)
info_log.place(x=50, y=350, width=200, height="73")
info_log.bind("<Key>", lambda e: handle_key_event(e))


if platform.system() == "Windows":
    # 윈도우 경로 설정 / Windows Path Configuration
    uPath = os.environ.get("UPATH", "勤怠管理_")
    bPath = os.environ.get("BPATH", ".\\backup\\勤怠管理_")
    rPath = os.environ.get("RPATH", ".\\resources\\")
    dePath = os.environ.get("DEPATH", ".\\EAR_default.xlsx")
    ePath = os.environ.get("EPATH", "勤怠管理_")
    logPath = os.environ.get("LOGPATH", ".\\backup\\SystemLog_")
    print("윈도우즈용 세팅 / Windows operating system")
else:
    # 맥 경로 설정 / Mac Path Configuration
    uPath = os.environ.get("UPATH", "勤怠管理_")
    bPath = os.environ.get("BPATH", "./backup/勤怠管理_")
    rPath = os.environ.get("RPATH", "./resources/")
    dePath = os.environ.get("DEPATH", "./resources/WTT_default.xlsx")
    ePath = os.environ.get("EPATH", "勤怠管理_")
    logPath = os.environ.get("LOGPATH", "./backup/SystemLog_")
    print("맥 및 기타 운영체제용 세팅 / Mac or other operating systems")

#admin password
# password = os.environ.get("AR_PASSWORD") #live
password = "1234" #테스트판 지정 패스워드    


#엑셀 파일 불러오기 및 시트 지정하기
wb = openpyxl.load_workbook(dePath)
ws = wb.worksheets[0]
ws_u = wb.worksheets[1]
today_start = ws.cell(row=ws.max_row, column=2).value
username_cell = ws.cell(row=1, column=2).value

#엑셀파일 유저명으로 재설정
if username_cell == "No_Username":
    wb = openpyxl.load_workbook(dePath)
    ws = wb.worksheets[0]

else :
    wb = openpyxl.load_workbook("{}{}.xlsx".format(uPath, username_cell))
    ws = wb.worksheets[0]   
    today_start = ws.cell(row=ws.max_row, column=2).value
    username_cell = ws.cell(row=1, column=2).value
    

#함수 정의(사용자 결정 버튼)
def handle_username_button():
    username_cell = ws.cell(row=1, column=2).value
    result_username =  input_box.get()
    result_username = result_username.replace(" ", "")
    result_username = result_username.replace("　", "")
    print("User name： "+result_username)
    
    if result_username == "":
        messagebox.showinfo("警告","名前が入力されていません。")
    else :
        label =create_custom_label("{}さん、ようこそ                       ".format(result_username))
        #시스템 로그
        log_text_file = open("{}{}.txt".format(logPath,result_username),'w')
        log_text_file.write(" ■名前：{}\n ■初回生成日：{} {}\n ______________________________\n".format(result_username,date_now,time_now))
        log_text_file.close()
        username_cell
        label.place(x=50, y=20)
        info_and_system_log("\n・アカウントを作成しました\n   アカウント名：{}\n   {} {}\n".format(result_username,date_now, time_now))

        #시스템 로그
        log_text_file = open("{}{}.txt".format(logPath,result_username),"a")
        log_text_file.write("\n・アカウントを作成しました\n   アカウント名：{}\n   {} {}\n".format(result_username,date_now, time_now))
        log_text_file.close()
        
        #유저명 로그 기록하기
        excel_username = ws.cell(row=1, column=2).value=result_username
        save_workbook_to_paths(uPath,result_username)
        wb.save(dePath)
        # ws.protection.password = os.environ.get("AR_EXCEL_PASSWORD") #live
        ws.protection.password = "1234" #테스트판 PW  
        ws.protection.enable() #시트보호하기

        #출/퇴근 버튼(초회 상단위치)
        create_button("勤務開始", handle_start_button, 50, 70, 200, 50)
        create_button("勤務終了", breaktime_type, 50, 130, 200, 50)

        #메뉴 활성
        handle_enable_menu()

        #유저 입력창 / 결정 버튼 비활성화 (생성 확정)
        username_button['state'] = tk.DISABLED
        input_box['state'] = tk.DISABLED
        send_line_message("\n[お知らせ]\n{}さんのアカウントが作成されました。\nアクションをお知らせします。", excel_username)

#유저 이름 취득
global excel_username 
excel_username = ws.cell(row=1, column=2).value  

global result_username


#함수 정의(출근 버튼)
def handle_start_button():
    global excel_username
    max_row_date = ws.cell(row=ws.max_row, column=1).value
    max_row_st = ws.cell(row=ws.max_row, column=2).value
    excel_username = ws.cell(row=1, column=2).value
    yd_end_null = ws.cell(row=ws.max_row, column=3).value
    target_col = "A"
    
    for row in range(1, ws.max_row+1):
        if date_now in str(ws[f"{target_col}{row}"].value):    
            send_line_message("\n[お知らせ]\n名前：{}\n\n出勤しているのに出勤ボタンを再度押しました。\nこのアクションはログに残しません。\n出勤時間：{}", excel_username, max_row_st)
            messagebox.showinfo("お知らせ","[お知らせ]\n本日の出勤データがすでにあります\n\nログを閲覧してください")
            info_and_system_log("\n・本日の出勤データがすでにあります\n   ログを閲覧してください\n   {} {}\n")
            return  # 중복 기록을 감지했을 때 함수 종료

    if max_row_date == date_now:
        send_line_message("\n[お知らせ]\n名前：{}\n\n出勤しているのに出勤ボタンを再度押しました。\nこのアクションはログに残しません。\n出勤時間：{}", excel_username, max_row_st)
        messagebox.showinfo("お知らせ","[お知らせ]\n本日はすでに出勤しています。\n出勤時間：{}\n\nこのアクションはログに残しません。".format(max_row_st))
        info_and_system_log("\n・すでに本日の出勤履歴があります\n   {} {}\n")
        return  # 중복 기록을 감지했을 때 함수 종료

    # 중복 기록이 아닌 경우 계속해서 기록을 수행
    j = ws.max_row+1
    today_date = ws.cell(row=j, column=1).value = date_now
    today_st = ws.cell(row=j, column=2).value = time_now

    send_line_message("\n[処理完了]\n名前：{}\n\n日付：{}\n出勤時間 : {}\n\n出勤時間をログに残しました。", excel_username, today_date, today_st)
    messagebox.showinfo("処理完了","[処理完了]\n[{}]\n出勤時間 : {}\n\n出勤時間をログに残しました。".format(today_date, today_st))
    info_and_system_log("\n・出勤時間を記録しました\n   {} {}\n")

    try:
        save_workbook_to_paths()
    except:
        info_and_system_log()


#색상 설정
bg_var = StringVar()    
bg_var.set ("#323332")

fg_var = StringVar()    
fg_var.set ("#FFFFFF")

btn_bg_var = StringVar()    
btn_bg_var.set ("#0070C0")

#메뉴 표시 제어
var = StringVar()    
var.set ("disable")   

rdVer = tkinter.StringVar()
rdVer.set(4)
rd_label =create_custom_label("■ 休憩タイプを選択してください                   ")
radio_button1 = tkinter.Radiobutton(window, bg=bg_var.get(), fg="#FF4264", font=font, activebackground=bg_var.get(), activeforeground=fg_var.get(), value="01:00", variable=rdVer, text='1回休憩')
radio_button2 = tkinter.Radiobutton(window, bg=bg_var.get(), fg="#FF4264",font=font, activebackground=bg_var.get(), activeforeground=fg_var.get(), value="02:00", variable=rdVer,  text='2回休憩')
radio_button3 = tkinter.Radiobutton(window,bg=bg_var.get(), fg="#FF4264",font=font, activebackground=bg_var.get(), activeforeground=fg_var.get(), value="00:00", variable=rdVer,  text='休憩なし')

start_button = tkinter.Button(window, highlightbackground="#323332", font=font, text="勤務開始")
end_button = tkinter.Button(window, highlightbackground="#323332", font=font, text="勤務終了")


def end_button_cancel_function():
    ws.cell(row=1, column=2).value
    handle_enable_menu()
    rd_label.place_forget()
    radio_button1.place_forget()
    radio_button2.place_forget()
    radio_button3.place_forget()
    rd_button_cancel.place_forget()
    rd_button.place_forget()
    rdVer.set(4)
    start_button.configure(state = 'normal')
    end_button.configure(state = 'normal')
    info_and_system_log("\n・勤務終了をキャンセルしました\n   {} {}\n")
    

def breaktime_type():
    ws = wb.worksheets[0]
    max_row_date = ws.cell(row=ws.max_row, column=1).value
    excel_username = ws.cell(row=1, column=2).value

    if max_row_date == date_now :
        disable_menu()
        start_button.configure(state = 'disabled')
        end_button.configure(state = 'disabled')
        
        #휴식시간 선택하기
        var.set(0)
        rd_label.place(x=50, y=190)
        radio_button1.place(x=50, y=220)
        radio_button2.place(x=50, y=240)
        radio_button3.place(x=50, y=260)
        rd_button_cancel.place(x=50,y=290, width=100, height=50)
        rd_button.place(x=150,y=290, width=100, height=50)     
        rd_result = rdVer.get()


    else :
        info_and_system_log("\n・本日の出勤記録がありません\n   {} {}\n")
        send_line_message("\n[お知らせ]\n名前：{}\n\n出勤していないのに退勤ボタンを押しました。\nこのアクションはログに残しません。", excel_username)
        messagebox.showinfo("警告","出勤履歴がないため、退勤することはできません。")


#함수 정의(퇴근 버튼)
def end_button_function():
    update_datetime()
    max_row_date = ws.cell(row=ws.max_row, column=1).value
    excel_username = ws.cell(row=1, column=2).value
        
    if max_row_date == date_now :
        rd_result = rdVer.get()
        j = ws.max_row     
        today_et = ws.cell(row=j, column=3).value=time_now
        rd_result = ws.cell(row=j, column=4).value=rd_result

        if rd_result == "4":
            messagebox.showinfo("お知らせ","[お知らせ]\n休憩タイプを選択してください")
            info_and_system_log("\n・休憩タイプを選択してください\n   {} {}\n")
            rdVer.set(4)

        else :    
            if rd_result == "01:00":
                ws.cell(row=j, column=4).value="01:00"

                try:
                    save_workbook_to_paths()
                except:
                    info_and_system_log()             
                
                send_line_message("\n[処理完了]\n名前：{}\n\n日付：{}\n退勤時間 : {}\n退勤時間をログに残しました。\n\n本日の退勤履歴がある場合は履歴が更新されます。",excel_username, date_now, today_et)
                messagebox.showinfo("処理完了","[処理完了]\n[{}]\n退勤時間 : {}\n退勤時間をログに残しました。\n\n本日の退勤履歴がある場合は履歴が更新されます。".format(date_now, today_et))
                info_and_system_log("\n・退勤時間を記録しました\n   {} {}\n")
                widgets_to_forget = [rd_label, radio_button1, radio_button2, radio_button3, rd_button_cancel, rd_button]

                for widget in widgets_to_forget:
                    widget.place_forget()
                
                rdVer.set(4)

                start_button.configure(state = 'normal')
                end_button.configure(state = 'normal')
                handle_enable_menu()             

            elif rd_result == "02:00":
                ws.cell(row=j, column=4).value="02:00"

                try:
                    save_workbook_to_paths()
                except:
                    info_and_system_log()
                    
                send_line_message("\n[処理完了]\n名前：{}\n\n日付：{}\n退勤時間 : {}\n退勤時間をログに残しました。\n\n本日の退勤履歴がある場合は履歴が更新されます。", excel_username, date_now, today_et)
                messagebox.showinfo("処理完了","[処理完了]\n[{}]\n退勤時間 : {}\n退勤時間をログに残しました。\n\n本日の退勤履歴がある場合は履歴が更新されます。".format(date_now, today_et))
                info_and_system_log("\n・退勤時間を記録しました\n   {} {}\n")
                widgets_to_remove = [rd_label, radio_button1, radio_button2, radio_button3, rd_button_cancel, rd_button]

                for widget in widgets_to_remove:
                    widget.place_forget()
                rdVer.set(4)

                start_button.configure(state = 'normal')
                end_button.configure(state = 'normal')
                handle_enable_menu()

            elif rd_result == "00:00":
                ws.cell(row=j, column=4).value="00:00"

                try:
                    save_workbook_to_paths()
                except:
                    info_and_system_log()
                    
                
                send_line_message("\n[処理完了]\n名前：{}\n\n日付：{}\n退勤時間 : {}\n退勤時間をログに残しました。\n\n本日の退勤履歴がある場合は履歴が更新されます。", excel_username, date_now, today_et)
                messagebox.showinfo("処理完了","[処理完了]\n[{}]\n退勤時間 : {}\n退勤時間をログに残しました。\n\n本日の退勤履歴がある場合は履歴が更新されます。".format(date_now, today_et))
                info_and_system_log("\n・退勤時間を記録しました\n   {} {}\n")
                
                widgets_to_forget = [rd_label, radio_button1, radio_button2, radio_button3, rd_button_cancel, rd_button]
                for widget in widgets_to_forget:
                    widget.place_forget()
                rdVer.set(4)

                start_button.configure(state = 'normal')
                end_button.configure(state = 'normal')
                handle_enable_menu()    
    else :
        info_and_system_log("\n・本日の出勤記録がありません\n   {} {}\n")
        send_line_message("\n[お知らせ]\n名前：{}\n\n出勤していないのに退勤ボタンを押しました。\nこのアクションはログに残しません。", excel_username)
        messagebox.showinfo("警告","出勤履歴がないため、退勤することはできません。まず、業務開始ボタンを押してください。")
    

rd_button_cancel = create_button("キャンセル", end_button_cancel_function)
rd_button = create_button("決定", end_button_function)

#메인 로고
# shion_logo = PhotoImage(file = "{}shion_logo.png".format(rPath))
# shion_logo = PhotoImage(file = "{}shion_logo.png".format(rPath))
# shion_cr = Label(window, bg="#323332", image = shion_logo)
shion_cr = Label(window, bg="#323332", text="© 2023 Shion")
shion_cr.place(x=100, y=430)

#메뉴바
menubar = tk.Menu(window)
window.config(menu = menubar)

item3 = tk.Menu(menubar,tearoff=0)
menubar.add_cascade(label="File", menu =item3) #File메뉴바

item1 = tk.Menu(menubar,tearoff=0)
menubar.add_cascade(label="User", menu =item1) #User메뉴바

item = tk.Menu(menubar,tearoff=0)
menubar.add_cascade(label="Admin", menu =item) #Admin메뉴바

item2 = tk.Menu(menubar,tearoff=0)
menubar.add_cascade(label="Help", menu =item2) #Help메뉴바


#메뉴 비활성화 함수 정의
def disable_menu():
    item.entryconfigure(0,state="disable") 
    item.entryconfigure(1,state="disable") 
    item.entryconfigure(3,state="disable")
    item.entryconfigure(4,state="disable")
    item.entryconfigure(6,state="disable")
    
    item1.entryconfigure(0,state="disable")
    item1.entryconfigure(1,state="disable")
    item1.entryconfigure(3,state="disable")
    item1.entryconfigure(4,state="disable")
    item1.entryconfigure(6,state="disable")
    item1.entryconfigure(7,state="disable")
    
    item3.entryconfigure(0,state="disable")

#메뉴 활성화 함수 정의
def handle_enable_menu():
    item.entryconfigure(0,state="normal") 
    item.entryconfigure(1,state="normal") 
    item.entryconfigure(3,state="normal")
    item.entryconfigure(4,state="normal")
    item.entryconfigure(6,state="normal")

    item1.entryconfigure(0,state="normal")     
    item1.entryconfigure(1,state="normal")
    item1.entryconfigure(3,state="normal")
    item1.entryconfigure(4,state="normal")
    item1.entryconfigure(6,state="normal")
    item1.entryconfigure(7,state="normal")
                                        
    item2.entryconfigure(0,state="normal")

    item3.entryconfigure(0,state="normal")


#함수 정의(시스템 로그창)
def sys_win():

    #서브(시스템로그)
    sys_window = tkinter.Tk() 
    sys_window.title("SYSTEM LOG") 
    sys_window.geometry("300x470") 
    sys_window.resizable(0,0) #
    sys_window.configure(bg='#323332')
    set_font(sys_window)

    write_system_log = tkinter.Text(sys_window, highlightbackground="#323332", bg="#1E1E1E", fg="#ffffff", font=font_log)  #font=font_log 
    write_system_log.place(x=20, y=20, width=260, height=400)
    write_system_log.bind("<Key>", lambda e: handle_key_event(e))
    
    #시스템 로그 기록
    excel_username = ws.cell(row=1, column=2).value
    system_text_file = open("{}{}.txt".format(logPath,excel_username),"r")
    system_text_file_read = system_text_file.read()
    write_system_log.insert(tkinter.CURRENT, system_text_file_read)

    #찾기 기능 UI
    sys_log_label = tkinter.Label(sys_window, bg=bg_var.get(), fg=fg_var.get(), font=font, text="検索 : ")
    sys_log_label.place(x=20, y=431)

    edit = tkinter.Entry(sys_window, bg="#1E1E1E", font=font, fg="#ffffff", highlightbackground="#323332",insertbackground='#ffffff')
    edit.place(x=60, y=428, width=160, height=32)

    sys_log_btn = tkinter.Button(sys_window, highlightbackground="#323332", font=font, text="実行")
    sys_log_btn.place(x=230, y=428, width=50, height=32)


    #함수 정의 (시스템 로그 찾기)
    def find():
        write_system_log.tag_remove('found', '1.0', END)
        s = edit.get()
        try:
            if s:
                idx = '1.0'
                while 1:
                    idx = write_system_log.search(s, idx, nocase=1,
                                    stopindex=END)
                    if not idx: 
                        break

                    lastidx = '%s+%dc' % (idx, len(s))
                    write_system_log.tag_add('found', idx, lastidx)
                    idx = lastidx
                write_system_log.tag_config('found', foreground='red')
                print(lastidx) #고의로 에러를 발생 시킴  
            else:
                messagebox.showinfo("お知らせ","内容を入力してください")
        except:
            messagebox.showinfo("お知らせ","該当する内容がありません")
        edit.focus_set()
    sys_log_btn.config(command=find)
    sys_window.mainloop()


#메뉴바 함수정의
def admin_insert_start(): 
    excel_username = ws.cell(row=1, column=2).value
    
    #관리자 모드 안내
    create_custom_label("[管理者モード] 開始時間の変更                   ", 50, 20)
    
    #출근 / 퇴근 버튼 비활성화
    def admin_button1_function() :
        input1_admin.focus()
        admin_font = Font(color="00FF0000")

        target_string = input1_admin.get() #"수동 날짜 취득"
        insert_string = input2_admin.get() #시간 취득

        target_string = target_string.translate(str.maketrans({chr(0xFF01 + i): chr(0x21 + i) for i in range(94)}))
        insert_string = insert_string.translate(str.maketrans({chr(0xFF01 + i): chr(0x21 + i) for i in range(94)}))

        input1_admin.delete(0, tk.END)
        input1_admin.insert(0,target_string)

        input2_admin.delete(0, tk.END)
        input2_admin.insert(0,insert_string)        
        
        YorN = messagebox.askyesno("お知らせ","入力した内容を適用しますか")
        if YorN:            
            target_col = "A"
            next_col = "B"
            # nodata =""
            for row in range(1, ws.max_row+1):
                if target_string in str(ws[f"{target_col}{row}"].value):
                    ws[f"{next_col}{row}"] = insert_string
                    ws[f"{next_col}{row}"].font =  admin_font

                    try:
                        save_workbook_to_paths()
                    except:
                        info_and_system_log()
                        

                    input1_admin_button.configure(state = 'disable')
                    input1_admin_cancel_button.configure(state = 'disable')
                    input1_admin.configure(state = 'disable')
                    input2_admin.configure(state = 'disable')

                    messagebox.showinfo("処理完了","[処理完了]\n開始時間を修正しました")
                    
                    #하단 메뉴 자동으로 없애기
                    widgets_to_forget = [input1_admin, input2_admin, input1_admin_button, input1_admin_cancel_button, label_date, label_time]
                    for widget in widgets_to_forget:
                        widget.place_forget()
                        
                    handle_enable_menu()
                    window.after(1500,handle_timecheck)
                    info_and_system_log("\n・開始時間を修正しました\n   {} {}(修正)\n   {} {}\n".format(target_string, insert_string, date_now, time_now))
                    has_data = True
            
            if not has_data:
                messagebox.showinfo("お知らせ","[処理失敗]\n該当する日付のデータがありません。")
        else:
            info_and_system_log("\n・時間変更をキャンセルしました\n   {} {}\n") 
  

    #어드민 날짜/시간 입력 항목의 취소 (화면에서 없애기)    
    def admin_button1_cancel_function():
        ws.cell(row=1, column=2).value
        info_and_system_log("\n・修正モードをキャンセルしました\n  メニューをアンロックします\n   {} {}\n")

        #하단 메뉴 취소 버튼으로 없애기
        widgets_to_forget = [input1_admin, input2_admin, input1_admin_button, input1_admin_cancel_button, label_date, label_time]
        for widget in widgets_to_forget:
            widget.place_forget()

        admin_pw.configure(state = 'normal')
        admin_pw_button.configure(state = 'normal')
        admin_pw.delete(0,20)

        #메뉴 활성화
        handle_enable_menu()


    def admin_button2_function(): #패스워드창
        input_pw = admin_pw.get() #입력 패스워드
        excel_username = ws.cell(row=1, column=2).value

        if password == input_pw :        
            admin_pw.configure(state = 'disabled')
            admin_pw_button.configure(state = 'disabled')
            info_and_system_log("\n・修正モードに入りました\n   メニューをロックします\n   {} {}\n")

            #메뉴 비활성화
            disable_menu()

            #어드민 누락 입력창 만들기
            label_date.place(x=50, y=200)
            input1_admin.place(x=50, y=220,  width=100, height=50) 
            label_time.place(x=150, y=200)
            input2_admin.place(x=150, y=220,  width=100, height=50) 

            #어드민 누락 결정버튼 만들기
            input1_admin_cancel_button.place(x=50, y=280, width=100, height=50)
            input1_admin_button.place(x=150, y=280, width=100, height=50)
            try:
                save_workbook_to_paths()
            except:
                info_and_system_log()
                
            
        else :
            messagebox.showinfo("お知らせ","[お知らせ]\nパスワードが入力されていないか正しくありません。もう一度入力してください。")
            info_and_system_log("\n・パスワードが正しくありません\n   {} {}\n")

        try:       
            save_workbook_to_paths()
        except:
            info_and_system_log()
            

    #어드민 패스워드창 만들기
    create_custom_label("パスワード                                              ")
    create_admin_pw()
    info_and_system_log("\n・管理者モード(1)に入りました\n   {} {}\n")
    
    #어드민 패스워드 결정버튼 만들기
    admin_pw_button = tkinter.Button(window, highlightbackground="#323332", bg=btn_bg_var.get(), font=font, text="確認", command=admin_button2_function)
    admin_pw_button.place(x=50, y=130, width=200, height=50)
    admin_pw_button.config(bg='#F0F8FF')
    
    #어드민 날짜시간 입력 창 정의 (표시는 별도)
    label_date =create_custom_label("日付",)
    label_time =create_custom_label("時間") 
    input2_admin = create_custom_entry(font)
    input1_admin = create_custom_entry(font)
    input2_admin.insert(0,time_now)
    input1_admin.insert(0,date_now)

    #어드민 누락 결정버튼 정의 (표시는 별도)
    input1_admin_cancel_button = create_button("キャンセル", admin_button1_cancel_function)
    input1_admin_button = create_button("決定", admin_button1_function)

    try:    
        save_workbook_to_paths()  
    except:
        info_and_system_log()
        
    
#유급 관련 메뉴

#유급 신청 UI
def day_off():
    ws_u = wb.worksheets[1]
    ws.cell(row=1, column=2).value

    def day_off_apply():
        excel_username = ws.cell(row=1, column=2).value
        day_off_date = input_day_off_date.get()
        type_1_resule = combobox_type1.get()
        type_2_resule = combobox_type2.get()
        day_off_date = day_off_date.translate(str.maketrans({chr(0xFF01 + i): chr(0x21 + i) for i in range(94)}))

        if validate_date(day_off_date):
            input_day_off_date.delete(0, tk.END)
            input_day_off_date.insert(0,day_off_date)
            YorN = messagebox.askyesno("お知らせ","入力した内容で申請しますか")

            if YorN == True:
                day_off_date = input_day_off_date.get()
                type_1_resule = combobox_type1.get()
                type_2_resule = combobox_type2.get()

                ws_u.cell(row=ws_u.max_row, column=1).value
                j = ws_u.max_row+1
                # today_date = ws_u.cell(row=j, column=1).value=date_now
                day_off_date = ws_u.cell(row=j, column=2).value=day_off_date
                type_1_resule = ws_u.cell(row=j, column=3).value=type_1_resule

                if type_2_resule == "全休":
                    # type_2_resule_int = ws_u.cell(row=j, column=4).value=1.0
                    ws_u.cell(row=j, column=5).value=type_2_resule
                else:
                    # type_2_resule_int = ws_u.cell(row=j, column=4).value=0.5
                    ws_u.cell(row=j, column=5).value=type_2_resule

                #저장하기
                try:                
                    save_workbook_to_paths()
                except:
                    info_and_system_log()

                messagebox.showinfo("処理完了","[処理完了]\n有休を申請しました")
                info_and_system_log("\n・有休を申請しました\n   {} {}/{}(申請)\n   {} {}\n".format(day_off_date, type_1_resule, type_2_resule,date_now, time_now))
                send_line_message("\n[お知らせ]\n名前：{}\n\n申請日：{}\n有休日:{}\nタイプ：{}/{}", excel_username, date_now, day_off_date, type_1_resule, type_2_resule)
            else:
                info_and_system_log("\n・有休申請をキャンセルしました\n   {} {}\n")

        else:
            messagebox.showerror("エラー","日付が正しくありません\nもう一度入力してください")
            info_and_system_log("\n・日付が正しくありません\n   もう一度入力してください\n   {} {}\n")
    
    create_custom_label("[有休申請] 承認後に申請してください                       ", 50, 20)

    label1 = tkinter.Label(window, bg=bg_var.get(), fg="orange", font=font, text="メールまたは口頭での上長承認後に\n申請してください")
    label1.place(x=50, y=250)
    label1.after(6000, label1.place_forget)
    create_custom_label("日付           ", 50, 50)
    create_custom_label("タイプ         ", 50, 50)
    input_day_off_date = tkinter.Entry(window, bg="#1E1E1E", fg="#ffffff",highlightbackground="#323332", insertbackground='#ffffff', font=font) #전체 UI에 맞는 설정
    input_day_off_date.place(x=50, y=70,  width=100, height=50)
    input_day_off_date.insert(0,date_now)
    input_day_off_date.focus
    info_and_system_log("\n・有休申請モードに入りました\n   {} {}\n")

    #combobox
    type_list1 = ("事前", "事後")
    type_list2 = ("全休","午前半休","午後半休")
    combobox_type1 = ttk.Combobox(window,values=type_list1, state="readonly")
    combobox_type1.place(x=150,y=70, width=100, height=25)
    combobox_type1.set("事前")
    combobox_type2 = ttk.Combobox(window, values=type_list2, state="readonly")
    combobox_type2.place(x=150,y=95, width=100, height=25)
    combobox_type2.set("全休")

    #결정 버튼
    end_button = create_button("申請", day_off_apply, 50, 130, 200, 50)


#수동 변경 관련 (신규 추가)
def add_new_record (): 
    excel_username = ws.cell(row=1, column=2).value
    
    #관리자 모드 안내
    create_custom_label("[管理者モード] 出勤/退勤新規生成                       ", 50, 20)
    
    #여기에 기능 정의하기
    def admin_button1_function() :
        ws = wb.worksheets[0]       
        admin_font = Font(color="00FF0000")       
        target_string = input1_admin.get() 
        insert_start_string = input2_admin.get() 
        insert_end_string = input3_admin.get() 
        rd_result = rdVer.get() 

        target_string = target_string.translate(str.maketrans({chr(0xFF01 + i): chr(0x21 + i) for i in range(94)}))
        insert_start_string = insert_start_string.translate(str.maketrans({chr(0xFF01 + i): chr(0x21 + i) for i in range(94)}))
        insert_end_string = insert_end_string.translate(str.maketrans({chr(0xFF01 + i): chr(0x21 + i) for i in range(94)}))

        input1_admin.delete(0, tk.END)
        input1_admin.insert(0,target_string)

        input2_admin.delete(0, tk.END)
        input2_admin.insert(0,insert_start_string)

        input3_admin.delete(0, tk.END)
        input3_admin.insert(0,insert_end_string)

        target_col = "A"
        YorN =""

        for row in range(3, ws.max_row+1):
            #입력한 날짜가 이미 데이터로 남아 있는지 체크
            if target_string in str(ws[f"{target_col}{row}"].value):
                YorN = False
        
        #누락분 신규 추가 기능 정의하기
        if YorN:
            messagebox.showinfo("お知らせ","すでにデータがあります\nデータ修正モードをご利用ください")
            info_and_system_log("\n・すでにデータがあります\n   データ修正モードをご利用ください\n   {} {}\n")

        elif target_string >= date_now:
            messagebox.showinfo("お知らせ","[お知らせ]\n本日を含む未来は登録できません\n過去の日付を記載してください")
            log_text_file = open("{}{}.txt".format(logPath,excel_username),"a")
            log_text_file.write("\n・本日・未来は登録できません\n   過去の日付を記載してください\n   {} {}\n".format(date_now, time_now))
            log_text_file.close()

        elif rd_result == "4":
            messagebox.showinfo("お知らせ","[お知らせ]\n休憩タイプを選択してください")
            info_and_system_log("\n・休憩タイプを選択してください\n   {} {}\n")
            rdVer.set(4)

        elif validate_date(target_string) == False:
            messagebox.showerror("エラー","日付が正しくありません\nもう一度入力してください")
            info_and_system_log("\n・日付が正しくありません\n   もう一度入力してください\n   {} {}\n")

        else :
            YorN = messagebox.askyesno("お知らせ","入力した内容を適用しますか")
            if YorN :
                target_col = "A"
                minus_days = []
                str_datetime = target_string
                format = '%Y/%m/%d'
                dt_datetime = datetime.strptime(str_datetime,format)

                for i in range(32):
                    minus_days.append((dt_datetime - timedelta(days=i)).strftime("%Y/%m/%d"))

                def minus_days_1():
                    insert_row = row+1
                    ws.insert_rows(insert_row)
                    ws.cell(row=insert_row, column=1).value = target_string
                    ws.cell(row=insert_row, column=2).value = insert_start_string
                    ws.cell(row=insert_row, column=3).value = insert_end_string
                    ws.cell(row=insert_row, column=1).font =  admin_font
                    ws.cell(row=insert_row, column=2).font =  admin_font
                    ws.cell(row=insert_row, column=3).font =  admin_font

                    #라디오 버튼 결과 데이터 추가
                    rd_result_mapping = {"01:00": "01:00", "02:00": "02:00", "00:00": "00:00"}
                    ws.cell(row=insert_row, column=4).value = rd_result_mapping.get(rd_result)

                    try:
                        save_workbook_to_paths()
                    except:
                        info_and_system_log()

                    input1_admin_cancel_button.configure(state = 'disabled')
                    input1_admin_button.configure(state = 'disabled')
                    input1_admin.configure(state = 'disabled')
                    input2_admin.configure(state = 'disabled')
                    input3_admin.configure(state = 'disabled')

                    messagebox.showinfo("処理完了","[処理完了]\n出勤/退勤時間を生成しました")
                    info_and_system_log("\n・出勤/退勤時間を生成しました\n   {} {} {}(a)\n   {} {}\n".format(target_string, insert_start_string, insert_end_string, date_now, time_now))
                    
                    #하단 메뉴 자동으로 없애기
                    widgets = [input1_admin, input2_admin, input3_admin, input1_admin_button, input1_admin_cancel_button, label_date, label_time_start, label_time_end, radio_button1, radio_button2, radio_button3]

                    for widget in widgets:
                        widget.place_forget()
                        
                    handle_enable_menu()
                    window.after(1500,handle_timecheck)
                    # write_to_system_log("\n・出勤/退勤時間を生成しました\n   {} {} {}(追加)\n   {} {}\n")

                for row in range(366, 0, -1):
                    for i in range(32):
                        if minus_days[i] in str(ws[f"{target_col}{row}"].value):
                            minus_days_1()
                            break
                    else:
                        continue
                    break        
            
            else:
                info_and_system_log("\n・出勤/退勤生成をキャンセルしました\n   {} {}\n")


    def admin_button2_function(): #패스워드창
        input_pw = admin_pw.get() #입력 패스워드

        if password == input_pw :
            info_and_system_log("\n・記録生成モードに入りました\n  メニューをロックします\n   {} {}\n")
            disable_menu()

            admin_pw.configure(state = 'disabled')
            admin_pw_button.configure(state = 'disabled')

            #어드민 누락 입력창 만들기 #####생성 관련
            label_date.place(x=50, y=190)
            input1_admin.place(x=50, y=210,  width=70, height=50)
            
            label_time_start.place(x=120, y=190)
            input2_admin.place(x=120, y=210,  width=65, height=50)

            label_time_end.place(x=185, y=190)
            input3_admin.place(x=185, y=210,  width=65, height=50)
            

            #라디오 버튼
            rdVer.set(4)
            radio_button1.place(x=50, y=270)
            radio_button2.place(x=120, y=270)
            radio_button3.place(x=190, y=270)
            radio_button3.configure(text='なし')


            #어드민 누락 결정버튼 만들기 #####생성 관련
            input1_admin_cancel_button.place(x=50, y=300, width=100, height=50)
            input1_admin_button.place(x=150, y=300, width=100, height=50)      
            
            try:             
                save_workbook_to_paths() 
            except:
                info_and_system_log()
                

        else :
            messagebox.showinfo("お知らせ","[お知らせ]\nパスワードが入力されていないか正しくありません。もう一度入力してください。")
            info_and_system_log("\n・パスワードが正しくありません\n   {} {}\n")

        try:        
            save_workbook_to_paths()
        except:
            info_and_system_log()
            


    #어드민 날짜/시간 입력 항목의 취소 (화면에서 없애기)    
    def admin_button1_cancel_function():
        ws.cell(row=1, column=2).value       
        rdVer.set(4)
        info_and_system_log("\n・生成モードをキャンセルしました\n  メニューをアンロックします\n   {} {}\n")

        #하단 메뉴 취소 버튼으로 없애기
        widgets_to_remove = [input1_admin, input2_admin, input3_admin, input1_admin_button, input1_admin_cancel_button, label_date, label_time_start, label_time_end, radio_button1, radio_button2, radio_button3]

        for widget in widgets_to_remove:
            widget.place_forget()

        admin_pw.configure(state = 'normal')
        admin_pw_button.configure(state = 'normal')
        admin_pw.delete(0,20)

        #메뉴 활성화
        handle_enable_menu()


    #어드민 패스워드창 만들기
    create_custom_label("パスワード                               ")
    create_admin_pw()
    info_and_system_log("\n・管理者モード(3)に入りました\n   {} {}\n")

    #어드민 패스워드 결정버튼 만들기
    admin_pw_button = create_button("確認", admin_button2_function, 50, 130, 200, 50)

    #어드민 누락 입력창 정의 (표시 별도)
    label_date =create_custom_label("日付")
    label_time_start =create_custom_label("出勤時間")
    label_time_end =create_custom_label("退勤時間")
    input1_admin = create_custom_entry(font_log)
    input2_admin = create_custom_entry(font_log)
    input3_admin = create_custom_entry(font_log)
    input1_admin_cancel_button = create_button("キャンセル", admin_button1_cancel_function)
    input1_admin_button = create_button("決定", admin_button1_function)

    input1_admin.insert(0,date_now)
    input2_admin.insert(0,time_now)
    input3_admin.insert(0,time_now)


#수동 변경 관련 (한줄 삭제 추가)
def del_record():
    excel_username = ws.cell(row=1, column=2).value
    
    #관리자 모드 안내
    create_custom_label("[管理者モード] 出勤/退勤削除                            ", 50, 20)
    
    #여기에 기능 정의하기
    def admin_button1_function() :
        update_datetime() 
        
        target_string = input1_admin.get() #"수동 날짜 취득"
        target_string = target_string.translate(str.maketrans({chr(0xFF01 + i): chr(0x21 + i) for i in range(94)}))

        input1_admin.delete(0, tk.END)
        input1_admin.insert(0,target_string)

        target_col = "A"

        YorN = messagebox.askyesno("お知らせ","入力した内容を適用しますか")
        if YorN:
            target_col = "A"
            # next_col = "B"
            
            # nodata =""
            for row in range(1, ws.max_row+1):
                if target_string in str(ws[f"{target_col}{row}"].value):
                    ws.delete_rows(row)

                    try:
                        save_workbook_to_paths()
                    except:
                        info_and_system_log()                     

                    input1_admin_button.configure(state = 'disable')
                    input1_admin_cancel_button.configure(state = 'disable')
                    input1_admin.configure(state = 'disable')
                    input2_admin.configure(state = 'disable')
                    messagebox.showinfo("処理完了","[処理完了]\nデータを削除しました")
                    
                    #하단 메뉴 자동으로 없애기
                    widgets_to_forget = (input1_admin, input2_admin, input1_admin_button, input1_admin_cancel_button, label_date)

                    for widget in widgets_to_forget:
                        widget.place_forget()

                    handle_enable_menu()
                    window.after(1500,handle_timecheck)
                    info_and_system_log("\n・データを削除しました\n   {}(削除)\n   {} {}\n".format(target_string, date_now, time_now))
                    has_data = True
            
            if not has_data:
                messagebox.showinfo("お知らせ","[処理失敗]\n該当する日付のデータがありません。")

        else:
            info_and_system_log("\n・データ削除をキャンセルしました\n   {} {}\n") 


    def admin_button2_function(): #패스워드창
        input_pw = admin_pw.get() #입력 패스워드

        if password == input_pw :
            info_and_system_log("\n・データ削除モードに入りました\n  メニューをロックします\n   {} {}\n")

            #메뉴 비활성화
            disable_menu()

            admin_pw.configure(state = 'disabled')
            admin_pw_button.configure(state = 'disabled')

            #어드민 누락 입력창 만들기
            label_date.place(x=50, y=200)
            input1_admin.place(x=50, y=220,  width=200, height=50)

            #어드민 누락 결정버튼 만들기
            input1_admin_cancel_button.place(x=50, y=280, width=100, height=50)
            input1_admin_button.place(x=150, y=280, width=100, height=50)     
            
            try:             
                save_workbook_to_paths()
            except:
                info_and_system_log()            


        else :
            messagebox.showinfo("お知らせ","[お知らせ]\nパスワードが入力されていないか正しくありません。もう一度入力してください。")
            info_and_system_log("\n・パスワードが正しくありません\n   {} {}\n")

        try:        
            save_workbook_to_paths()
        except:
            info_and_system_log()          


    #어드민 날짜/시간 입력 항목의 취소 (화면에서 없애기)    
    def admin_button1_cancel_function():
        ws.cell(row=1, column=2).value
        info_and_system_log("\n・削除モードをキャンセルしました\n  メニューをアンロックします\n   {} {}\n")

        #하단 메뉴 취소 버튼으로 없애기
        widgets_to_forget = [input1_admin, input2_admin, input3_admin, input1_admin_button, input1_admin_cancel_button, label_date, label_time_start, label_time_end]

        for widget in widgets_to_forget:
            widget.place_forget()

        admin_pw.configure(state = 'normal')
        admin_pw_button.configure(state = 'normal')
        admin_pw.delete(0,20)

        #메뉴 활성화
        handle_enable_menu()


    #어드민 패스워드창 만들기
    label_pw =create_custom_label("パスワード                               ", 50, 50)
    create_admin_pw()
    info_and_system_log("\n・管理者モード(4)に入りました\n   {} {}\n")

    #어드민 패스워드 결정버튼 만들기
    admin_pw_button = create_button("確認", admin_button2_function, 50, 130, 200, 50)
    
    #어드민 누락 입력창 정의 (표시 별도)
    label_date =create_custom_label("日付")
    label_time_start =create_custom_label("出勤時間")
    label_time_end =create_custom_label("退勤時間")
    input1_admin = create_custom_entry(font_log)
    input2_admin = create_custom_entry(font_log)
    input3_admin = create_custom_entry(font_log)
    input1_admin_cancel_button = create_button("キャンセル", admin_button1_cancel_function)
    input1_admin_button = create_button("決定", admin_button1_function)

    input1_admin.insert(0,date_now)
    input2_admin.insert(0,time_now)
    input3_admin.insert(0,time_now)


#수동 변경 관련
def admin_insert_end():
    excel_username = ws.cell(row=1, column=2).value
    
    #관리자 모드 안내
    create_custom_label("[管理者モード] 終了時間の変更                   ", 50, 20)
    
    #여기에 기능 정의하기
    def admin_button1_function() :
        admin_font = Font(color="00FF0000")

        target_string = input1_admin.get() #"수동 날짜 취득"
        insert_string = input2_admin.get() #시간 취득

        target_string = target_string.translate(str.maketrans({chr(0xFF01 + i): chr(0x21 + i) for i in range(94)}))
        insert_string = insert_string.translate(str.maketrans({chr(0xFF01 + i): chr(0x21 + i) for i in range(94)}))

        input1_admin.delete(0, tk.END)
        input1_admin.insert(0,target_string)

        input2_admin.delete(0, tk.END)
        input2_admin.insert(0,insert_string)
        
        YorN = messagebox.askyesno("お知らせ","入力した内容を適用しますか")
        if YorN:
            target_col = "A"
            next_col = "C"

            # nodata =""
            for row in range(3, ws.max_row+1):
                if target_string in str(ws[f"{target_col}{row}"].value):
                    ws[f"{next_col}{row}"] = insert_string
                    ws[f"{next_col}{row}"].font =  admin_font

                    try:
                        save_workbook_to_paths()
                    except:
                        info_and_system_log()
                        
                    input1_admin_cancel_button.configure(state = 'disabled')
                    input1_admin_button.configure(state = 'disabled')
                    input1_admin.configure(state = 'disabled')
                    input2_admin.configure(state = 'disabled')
                    messagebox.showinfo("処理完了","[処理完了]\n終了時間を修正しました")
                    
                    #하단 메뉴 자동으로 없애기
                    for widget in [input1_admin, input2_admin, input1_admin_button, input1_admin_cancel_button, label_date, label_time]:
                        widget.place_forget()
                    handle_enable_menu()

                    window.after(1500,handle_timecheck)
                    info_and_system_log("\n・終了時間を修正しました\n   {} {}(修正)\n   {} {}\n".format(target_string, insert_string, date_now, time_now))
                    has_data = True

            if not has_data :
                messagebox.showinfo("お知らせ","[処理失敗]\n該当する日付のデータがありません。")
        
        else:
            info_and_system_log("\n・時間変更をキャンセルしました\n   {} {}\n")
            log_text_file = open("{}{}.txt".format(logPath,excel_username),"a")
            log_text_file.write("\n・時間変更をキャンセルしました\n   {} {}\n")
            log_text_file.close()


    def admin_button2_function(): #패스워드창
        input_pw = admin_pw.get() #입력 패스워드

        if password == input_pw :
            info_and_system_log("\n・修正モードに入りました\n  メニューをロックします\n   {} {}\n")
            disable_menu()

            admin_pw.configure(state = 'disabled')
            admin_pw_button.configure(state = 'disabled')

            #어드민 누락 입력창 만들기
            label_date.place(x=50, y=200)         
            input1_admin.place(x=50, y=220,  width=100, height=50)
            label_time.place(x=150, y=200)
            input2_admin.place(x=150, y=220,  width=100, height=50)

            #어드민 누락 결정버튼 만들기
            input1_admin_cancel_button.place(x=50, y=280, width=100, height=50)
            input1_admin_button.place(x=150, y=280, width=100, height=50)

            try:            
                save_workbook_to_paths()
            except:
                info_and_system_log()
                
        else :
            messagebox.showinfo("お知らせ","[お知らせ]\nパスワードが入力されていないか正しくありません。もう一度入力してください。")
            info_and_system_log("\n・パスワードが正しくありません\n   {} {}\n")

        try:       
            save_workbook_to_paths()
        except:
            info_and_system_log()
            

    #어드민 날짜/시간 입력 항목의 취소 (화면에서 없애기)    
    def admin_button1_cancel_function():
        ws.cell(row=1, column=2).value
        info_and_system_log("\n・修正モードをキャンセルしました\n  メニューをアンロックします\n   {} {}\n")

        #하단 메뉴 취소 버튼으로 없애기
        widgets_to_forget = (input1_admin, input2_admin, input1_admin_button, input1_admin_cancel_button, label_date, label_time)

        for widget in widgets_to_forget:
            widget.place_forget()

        admin_pw.configure(state = 'normal')
        admin_pw_button.configure(state = 'normal')
        admin_pw.delete(0,20)

        #메뉴 활성화
        handle_enable_menu()
    
    #어드민 패스워드창 만들기
    create_custom_label("パスワード                               ", 50, 50)
    create_admin_pw()
    info_and_system_log("\n・管理者モード(2)に入りました\n   {} {}\n")

    #어드민 패스워드 결정버튼 만들기
    admin_pw_button = create_button("確認", admin_button2_function, 50, 130, 200, 50)
#     admin_pw_button.configure(state = 'disabled')   

    #어드민 누락 입력창 정의 (표시 별도)
    label_date =create_custom_label("日付")
    label_time =create_custom_label("時間")
    input1_admin = create_custom_entry(font)
    input2_admin = create_custom_entry(font)
    input1_admin_cancel_button = create_button("キャンセル", admin_button1_cancel_function)
    input1_admin_button = create_button("決定", admin_button1_function)
    input2_admin.insert(0,time_now)
    input1_admin.insert(0,date_now)


#함수정의
def reset_user():
    ws.cell(row=1, column=2).value

    def reset_user_data(): #패스워드창
        input_pw = admin_pw.get() #입력 패스워드
        wb = openpyxl.load_workbook(dePath)
        ws = wb.worksheets[0]
        excel_username = ws.cell(row=1, column=2).value

        if password == input_pw :
            YorN = messagebox.askyesno("お知らせ","{}さんのアカウントを本当に初期化しますか".format(excel_username))
            
            if YorN:
                wb = openpyxl.load_workbook(dePath)
                ws = wb.worksheets[0]
                excel_username = ws.cell(row=1, column=2).value
                ws.cell(row=1, column=2).value="No_Username"
                wb.save(dePath)    
                admin_pw.configure(state = 'disabled')
                admin_pw_button.configure(state = 'disabled')
                disable_menu()

                #파일 삭제
                os.remove("{}{}.xlsx".format(uPath,excel_username))

                send_line_message("\n[お知らせ]\n名前：{}\n\nアカウントを初期化しました。", excel_username)
                messagebox.showinfo("処理完了","[処理完了]\n名前：{}\n\nアカウントを初期化しました\nアプリを終了してください".format(excel_username))
                info_and_system_log("\n・アカウントを初期化しました\n   アプリを終了してください\n   {} {}\n")


            else:
                admin_pw.delete(0,20)
                info_and_system_log("\n・初期化をキャンセルしました\n   {} {}\n")


        else :
            messagebox.showinfo("お知らせ","[お知らせ]\nパスワードが入力されていないか正しくありません。もう一度入力してください。")
            info_and_system_log("\n・パスワードが正しくありません\n   {} {}\n")
    

    #어드민 패스워드창 만들기 (유저 리셋 부분)
    
    #출근 / 퇴근 버튼 비활성화  
    create_custom_label("[管理者モード] アカウントの初期化                  ", 50, 20)    
    create_custom_label("パスワード                          ", 50, 50)
    create_admin_pw()
    info_and_system_log("\n・管理者モード(5)に入りました\n   {} {}\n")
    
    #어드민 패스워드 결정버튼 만들기
    admin_pw_button= create_button("確認", reset_user_data, 50,130, 200, 50)
    admin_pw_button.configure(state = 'normal')

#유저 이름 취득
excel_username = ws.cell(row=1, column=2).value


#User 메뉴 함수 처리
def send_msg_to_admin():
    excel_username = ws.cell(row=1, column=2).value

    def msg_button_function():
        user_msg = text_box.get('1.0', 'end')

        if user_msg == "\n":
            messagebox.showinfo("お知らせ","メッセージがありません\nメッセージを入力してください")
            info_and_system_log("\n・メッセージがありません\n   メッセージを入力してください\n   {} {}\n")

        else:
            YorN = messagebox.askyesno("お知らせ","管理者勤にメッセージを送信しますか")
            if YorN:
                messagebox.showinfo("お知らせ","管理者にメッセージを送信しました。")
                info_and_system_log("\n・メッセージを送信しました\n   {} {}\n")
                log_text_file = open(f"{logPath}{excel_username}.txt", "a")
                log_text_file.write(f"\n・メッセージを送信しました\n   メッセージ内容：\n   {user_msg}   \n   {date_now} {time_now}\n")
                log_text_file.close()
                send_line_message("\n[お知らせ]\n名前：{}\n\nメッセージ内容：\n{}", excel_username,user_msg)

                #유저 입력창 / 결정 버튼 비활성화 (생성 확정)
                text_box['state'] = tk.DISABLED
                msg_button['state'] = tk.DISABLED
                window.after(1500,handle_timecheck)

            else:
                info_and_system_log("\n・メッセージ送信をキャンセルしました\n   {} {}\n")

                log_text_file = open(f"{logPath}{excel_username}.txt", "a")
                log_text_file.write("\n・メッセージ送信をキャンセルしました\n   {} {}\n")
                log_text_file.close()


    #입력창 상단 라벨
    create_custom_label("メッセージ                              ")
    info_and_system_log("\n・メッセージ送信モードに入りました\n   {} {}\n")
    create_custom_label("管理者にメッセージを送信                          ", 50, 20)
    text_box = tkinter.Text(window, highlightbackground="#323332",bg="#1E1E1E", fg="#ffffff", insertbackground='#ffffff')
    text_box.place(x=50, y=70, width=200, height=50)
    text_box.focus()
    msg_button = create_button("送信", msg_button_function, 50, 130, 200, 50)


#관리자에게 근태 기록 수동으로 보내기 실행
def send_ear_to_admin_btn():
    excel_username = ws.cell(row=1, column=2).value

    YorN = messagebox.askyesno("お知らせ","管理者に怠記録を送信しますか")
    if YorN:
        send_time = datetime.now().strftime("%Y%m%d")
        try:
            wb.save("{}{}_{}.xlsx".format(bPath,excel_username,send_time)) ##파일서버 등으로 // 경로 바꾸기 필요
        except:
            info_and_system_log("\n・ネットワークの接続が不安定です\n   ネットワーク環境をご確認ください\n   {} {}\n")
            

        else:
            send_line_message("\n[処理完了]\n名前：{}\n\n勤怠記録が送信されました。\n共有フォルダをご確認ください。\nファイル名：勤怠管理_{}_{}.xlsx", excel_username,excel_username,send_time)
            messagebox.showinfo("お知らせ","勤怠記録を管理者に送信しました")
            info_and_system_log("\n・勤怠記録を送信しました\n   {} {}\n")

        window.after(2000,handle_timecheck)
    
    else:
        info_and_system_log("\n・送信をキャンセルしました\n   {} {}\n")


#관리자에게 근태 기록 수동으로 보내기 메뉴
def send_message_to_admin():
    excel_username = ws.cell(row=1, column=2).value
    info_and_system_log("\n・勤怠送信モードに入りました\n   {} {}\n")

    label = tkinter.Label(window, bg=bg_var.get(), fg=fg_var.get(), text="[勤怠記録] 現在までの勤怠記録を送信               ")
    label.place(x=50, y=20)
    create_custom_label("ファイル名                          ", 50, 50)
    send_time = datetime.now().strftime("%Y%m%d")
    start_button = tkinter.Label(window, highlightbackground="#323332", bg="#1E1E1E", fg=fg_var.get(), font=font_log, text="{}_{}.xlsx".format(excel_username,send_time)) #파일명 표시
    start_button.place(x=50, y=70, width=200, height=50)
    create_button("送信", send_ear_to_admin_btn, 50, 130, 200, 50)

def handle_timecheck():
    ws = wb.worksheets[0]
    excel_username = ws.cell(row=1, column=2).value
    label_time =create_custom_label("                                                   ")
    label_time.place(x=50, y=50)
    label = tkinter.Label(window, bg=bg_var.get(), fg=fg_var.get(), font = font, text=excel_username+"さん、おかえりなさい                         ")
    label.place(x=50, y=20) 
    
    #출/퇴근 버튼 만들기
    create_button("勤務開始", handle_start_button, 50, 70, 200, 50)
    create_button("勤務終了", breaktime_type, 50, 130, 200, 50)
    
    ws.cell(row=ws.max_row, column=3).value
    ws.cell(row=ws.max_row, column=1).value
    ws.cell(row=ws.max_row-1, column=3).value
    

def handle_search_log():
    excel_username = ws.cell(row=1, column=2).value

    def view_log_function() :
        df = pd.read_excel("{}{}.xlsx".format(uPath,excel_username), sheet_name="Raw_data", header=1)

        start_date = input_log_start.get() #시작 날짜 입력받기
        end_date = input_log_end.get() #종료 날짜 입력받기

        start_date = start_date.translate(str.maketrans({chr(0xFF01 + i): chr(0x21 + i) for i in range(94)}))
        end_date = end_date.translate(str.maketrans({chr(0xFF01 + i): chr(0x21 + i) for i in range(94)}))

        input_log_start.delete(0, tk.END)
        input_log_start.insert(0,start_date)

        input_log_end.delete(0, tk.END)
        input_log_end.insert(0,end_date)

        # df_result = df.query('"2022/12/27"<= date <= "2022/12/29"')
        df_result = df.query('"{}" <= Date <= "{}"'.format(start_date, end_date))

        #검색 조건의 해당 유무에 따른 처리
        if start_date == "" or end_date =="":
            messagebox.showinfo("お知らせ","[お知らせ]\n日付を入力してください")
            info_and_system_log("\n・日付を入力してください\n  {} {}\n")
        else:
            if np.all(df_result == ""):
                messagebox.showinfo("お知らせ","[お知らせ]\n該当するデータがありません")
                info_and_system_log("\n・該当するデータがありません\n   {} {}\n")
            else :
                #로그 안내창 만들기
                def handle_key_event(event):
                    if(12==event.state and event.keysym=='c' ):
                        return
                    else:
                        return "break"
                        
                info_log_search = tkinter.Text(window, highlightbackground="#323332", bg="#1E1E1E", fg="#ffffff",font=font_search)
                info_log_search.place(x=50, y=190, width=200, height="150")
                info_log_search.bind("<Key>", lambda e: handle_key_event(e))      
                info_log_search.insert(tkinter.CURRENT, df_result)
                info_and_system_log("\n・検索結果を閲覧しました\n  メニューはロック状態です\n   {} {}\n")
                disable_menu()

                #로그 초기화 버튼
                def handle_reset_log():                   
                    info_log_search.destroy()
                    reset_log_button.destroy()
                    handle_enable_menu()
                    info_and_system_log("\n・検索を終了しました\n  メニューをアンロックします\n   {} {}\n")

                reset_log_button = tkinter.Button(window, font=font, highlightbackground="#323332", text="検索終了", command=handle_reset_log)
                reset_log_button.place(x=50, y=130, width=200, height=50) #btn


    #로그 확인용 결정 버튼
    select_log_button = create_button("検索", view_log_function)

    #로그 검색 기능 표시
    date_now = datetime.now()
    before_one_week = date_now - timedelta(weeks=1)
    before_one_week_str = before_one_week.strftime("%Y/%m/%d")
    date_now_str = date_now.strftime("%Y/%m/%d")
    label_log_start = create_custom_label("開始          ")
    label_log_end = create_custom_label("終了           ") 
    input_log_start = create_custom_entry(font)
    input_log_end = create_custom_entry(font)
    input_log_start.insert(0, before_one_week_str)
    input_log_end.insert(0, date_now_str)
    input_log_start.focus()
    info_and_system_log("\n・ログ検索モード(1)に入りました\n   {} {}\n")
    create_custom_label("[ログ検索] 出勤・退勤記録                      ", 50, 20)

    label_log_start.place(x=50, y=50)
    label_log_end.place(x=150, y=50)
    select_log_button.place(x=50, y=130, width=200, height=50) #btn
    input_log_start.place(x=50, y=70,  width=100, height=50) #input start
    input_log_end.place(x=150, y=70,  width=100, height=50) #input end


#유급 신청 로그
def handle_vacation_log():
    excel_username = ws.cell(row=1, column=2).value

    def handle_search_vacation_log() :
        df = pd.read_excel("{}{}.xlsx".format(uPath,excel_username), sheet_name="有休履歴") #, header=1

        start_date = input_log_start.get() #시작 날짜 입력받기
        end_date = input_log_end.get() #종료 날짜 입력받기

        start_date = start_date.translate(str.maketrans({chr(0xFF01 + i): chr(0x21 + i) for i in range(94)}))
        end_date = end_date.translate(str.maketrans({chr(0xFF01 + i): chr(0x21 + i) for i in range(94)}))

        input_log_start.delete(0, tk.END)
        input_log_start.insert(0,start_date)

        input_log_end.delete(0, tk.END)
        input_log_end.insert(0,end_date)


        df1 = df.filter(items = ['有休日', '申請タイプ', '有休タイプ'])
        df_result = df1.query('"{}" <= 有休日 <= "{}"'.format(start_date, end_date))

        #검색 조건의 해당 유무에 따른 처리
        if start_date == "" or end_date =="":
            messagebox.showinfo("お知らせ","[お知らせ]\n日付を入力してください")
            info_and_system_log("\n・日付を入力してください\n  {} {}\n")

        else:
            if np.all(df_result == ""):
                messagebox.showinfo("お知らせ","[お知らせ]\n該当するデータがありません")
                info_and_system_log("\n・該当するデータがありません\n   {} {}\n")

            else :
                #유급 로그 안내창 만들기
                def handle_key_event(event):
                    if(12==event.state and event.keysym=='c' ):
                        return
                    else:
                        return "break"
                        
                info_log_search = tkinter.Text(window, highlightbackground="#323332", bg="#1E1E1E", fg="#ffffff",font=font_search)
                info_log_search.place(x=50, y=190, width=200, height="150")
                info_log_search.bind("<Key>", lambda e: handle_key_event(e))         
                info_log_search.insert(tkinter.CURRENT, df_result)
                info_and_system_log("\n・検索結果を閲覧しました\n  メニューはロック状態です\n   {} {}\n")
                disable_menu()

                #로그 초기화 버튼
                def handle_reset_log():
                    wb = openpyxl.load_workbook("{}{}.xlsx".format(uPath, username_cell))
                    ws = wb.worksheets[0]
                    info_log_search.destroy()
                    reset_log_button.destroy()
                    handle_enable_menu()
                    info_and_system_log("\n・検索を終了しました\n  メニューをアンロックします\n   {} {}\n")

                reset_log_button = tkinter.Button(window, font=font, highlightbackground="#323332", text="検索終了", command=handle_reset_log)
                reset_log_button.place(x=50, y=130, width=200, height=50) #btn


    #유급 로그 확인용 결정 버튼
    select_log_button = create_button("検索", handle_search_vacation_log)

    #유급 로그 검색 기능 표시
    date_now = datetime.now()
    before_one_week = date_now - timedelta(weeks=1)
    after_four_weeks = date_now + timedelta(weeks=4)
    before_one_week = before_one_week.strftime("%Y/%m/%d")
    after_four_weeks = after_four_weeks.strftime("%Y/%m/%d")
    label_log_start =create_custom_label("開始          ")
    label_log_end =create_custom_label("終了           ") 
    input_log_start = create_custom_entry(font)
    input_log_end = create_custom_entry(font)
    input_log_start.insert(0, before_one_week)
    # input_log_end.insert(0, date_now)
    input_log_end.insert(0, after_four_weeks)
    input_log_start.focus

    info_and_system_log("\n・ログ検索モード(2)に入りました\n   {} {}\n")
    create_custom_label("[ログ検索] 有休申請                                 ", 50, 20)

    label_log_start.place(x=50, y=50)
    label_log_end.place(x=150, y=50)
    select_log_button.place(x=50, y=130, width=200, height=50) #btn
    input_log_start.place(x=50, y=70,  width=100, height=50) #input start
    input_log_end.place(x=150, y=70,  width=100, height=50) #input end


#도움말 만들기
#버전 정보
def version_info():
    messagebox.showinfo("バージョン情報","勤怠管理ツール\n\nVersion : 1.0.1\nUpdate : 2023/01/30\n\n© 2023 Shion\nshion.jung@gmail.com")

#이용규약
# def terms_of_use():
#     messagebox.showinfo("利用規約","◯◯◯")

def write_system_log():
    log_text_file = open(f"{logPath}{excel_username}.txt","r")
    log_text_file.close()


#메뉴바  

#색상 설정
bg_var = StringVar()    
bg_var.set ("#323332")

fg_var = StringVar()    
fg_var.set ("#FFFFFF")

btn_bg_var = StringVar()    
btn_bg_var.set ("#0070C0")

#메뉴 표시 제어
var = StringVar()    
var.set ("disable")   

#메뉴바 File
item3.add_command(label="システムログ", command = sys_win, state = var.get(), font=font) #command = write_system_log
item3.add_separator()
item3.add_command(label="終了", command = window.destroy, font=font)

#메뉴바 User
item1.add_command(label="業務時間チェック", command = handle_timecheck, state = var.get(), font=font)
item1.add_command(label="出勤/退勤ログ検索", command = handle_search_log, state = var.get(), font=font)
item1.add_separator()
item1.add_command(label="管理者にメッセージを送信", command = send_msg_to_admin, state = var.get(), font=font)
item1.add_command(label="管理者に勤怠記録を送信", command = send_message_to_admin, state = var.get(), font=font)
item1.add_separator()
item1.add_command(label="有休申請", command = day_off, state = var.get(), font=font)
item1.add_command(label="有休申請ログ検索", command = handle_vacation_log, state = var.get(), font=font)


#메뉴바 Admin
item.add_command(label="開始時間変更", command = admin_insert_start, state = var.get(), font=font)
item.add_command(label="終了時間変更", command = admin_insert_end, state = var.get(), font=font)
item.add_separator()
item.add_command(label="出勤/退勤新規生成", command = add_new_record, state = var.get(), font=font)
item.add_command(label="出勤/退勤削除", command = del_record, state = var.get(), font=font)
item.add_separator()
item.add_command(label="アカウント初期化", command = reset_user, state = var.get(), font=font)

#메뉴바 About 
item2.add_command(label="バージョン", command = version_info, state ="normal", font=font)
item2.add_command(label="利用規約", state ="disabled", font=font)


#사용자 입력창 값 판별 (UI표시/비표시)
username_cell = ws.cell(row=1, column=2).value

if username_cell == "No_Username":
    
    label = tkinter.Label(window, bg=bg_var.get(), fg=fg_var.get(), font=font, text="名前を入力してください。")
    label.place(x=50, y=20)
    info_and_system_log("\n・初起動\n   {} {}\n")

    #입력창 만들기
    input_box = tkinter.Entry(window, bg="#1E1E1E", font=font, fg="#ffffff", highlightbackground="#323332",insertbackground='#ffffff')
    input_box.place(x=50, y=70, width=200, height=50) #인풋 박스가 표시될 좌표를 인수에 넣기
    input_box.focus() 

    #입력창 결정버튼 만들기
    username_button = create_button("決定", handle_username_button, 50, 130, 200, 50)


else :
    #메뉴 활성
    update_datetime()
    handle_enable_menu()

    excel_username = ws.cell(row=1, column=2).value
    label = tkinter.Label(window, bg=bg_var.get(), fg=fg_var.get(), font = font, text=excel_username+"さん、おかえりなさい")
    label.place(x=50, y=20)
    
    #출/퇴근 버튼 만들기
    start_button = create_button("勤務開始", handle_start_button, 50, 70, 200, 50)
    end_button = create_button("勤務終了", breaktime_type, 50, 130, 200, 50)
    
    yd_end_null = ws.cell(row=ws.max_row, column=3).value
    max_row_date = ws.cell(row=ws.max_row, column=1).value
    tdyd_null = ws.cell(row=ws.max_row-1, column=3).value

    #전일 퇴근을 깜박 있었을 때
    if max_row_date != date_now:   
        if yd_end_null is None :
            messagebox.showinfo("お知らせ","[お知らせ]\n前日の退勤記録がありません。\n管理者に報告してください。")
            info_and_system_log("\n・前日の退勤記録がありません\n   {} {}\n")
    else:
        if tdyd_null is None :
            info_and_system_log("\n・前日の退勤記録がありません\n   管理者に報告してください\n   {} {}\n")
            
window.mainloop()